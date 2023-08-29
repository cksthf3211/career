import openpyxl

# 생성
total_wb = openpyxl.Workbook()
# 시트 선택
total_ws = total_wb.active
# 시트 이름
total_ws.title = 'total'
# 헤더 추가
total_ws.append(['순번', '제품명', '가격', '수량', '합계'])

# 취합목록
save_path1 = 'Data/03.엑셀자동화/11번가.xlsx'
save_path2 = 'Data/03.엑셀자동화/스마트스토어.xlsx'
save_path3 = 'Data/03.엑셀자동화/쿠팡.xlsx'

file_list = ['11번가', '스마트스토어', '쿠팡']

for file in file_list:
    wb = openpyxl.load_workbook(f'Data/03.엑셀자동화/{file}.xlsx', data_only=True)
    ws = wb.active
    
    for row in ws.iter_rows(min_row = 2):
        data = []
        
        for cell in row:
            data.append(cell.value)
            
        total_ws.append(data)
        
# 순번 업데이트
for row in total_ws.iter_rows(min_row=2, max_col=1):
    for cell in row:
        cell.value = row[0].row - 1 # = row[0]은 2가 나옴 (min_row의 값)
      
# 순번 업데이트  
# i = 0
# for cell in total_ws['A']:
#     if i != 0:
#         cell.value = i
#     i = i + 1
    

total_wb.save('Data/03.엑셀자동화/total.xlsx')
    