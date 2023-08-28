import openpyxl

# 새로운 엑셀 파일 생성
wb = openpyxl.Workbook()

# 새로운 시트 생성
ws = wb.create_sheet('2030.01')

# 모든 시트 이름 출력
print(wb.sheetnames)

# 시트 삭제
del wb['Sheet']

wb.save('Data/03.엑셀자동화/자동화된엑셀.xlsx')