import openpyxl

save_path = 'Data/03.엑셀자동화/total.xlsx'

# 기존 엑셀 파일 불러오기
wb = openpyxl.load_workbook(save_path, data_only=True)
ws = wb.active

# 제품이름 리스트
name_list = []

for row in ws.iter_rows(min_row=2, min_col=2): # 2~마지막행, 2~마지막열
    name = row[0].value # 제품명
    # 시트 만듬
    if name not in name_list:
        # 리스트 돌면서 그 이름이 없으면 append and sheet create 함
        name_list.append(name)
        wb.create_sheet(name)
    
    data = []
    for cell in row:
        data.append(cell.value)
    wb[name].append(data)
        
wb.save(save_path)