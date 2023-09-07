import openpyxl

path = r'크롤링/Project/02.파이썬엑셀다루기/참가자_data.xlsx'

#새로운 엑셀 파일 생성
wb = openpyxl.load_workbook(path)

# 엑셀 시트선택
ws = wb['오징어게임']

# 데이터수정하기
ws['A3'] = '456'
ws['B3'] = '성기훈'

wb.save(path)

print('크롤링 작업을 마쳤습니다.')