import openpyxl

#새로운 엑셀 파일 생성
wb = openpyxl.Workbook()

# 현재 활성화된 시트 선택
ws = wb.create_sheet('오징어게임')

ws['A1'] = '참가번호'
ws['B1'] = '성명'

ws['A2'] = '1'
ws['B2'] = '오일남'

# 저장 -> r'' = 모든 문자를 문자형대로
wb.save(r'크롤링/Project/02.파이썬엑셀다루기/참가자_data.xlsx')

print('크롤링 작업을 마쳤습니다.')