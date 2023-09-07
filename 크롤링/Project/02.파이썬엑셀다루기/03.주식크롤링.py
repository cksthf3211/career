import openpyxl
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time

chrome_options = Options()
# 꺼짐 방지 옵션 추가함
chrome_options.add_experimental_option("detach", True)
# 크롬창 안뜨게 함
chrome_options.add_argument('--headless') # headless 모드 활성화
chrome_options.add_argument('--disable-gpu') # GPU 가속 비활성화
# 쓸데없는 오류메세지같은 이상한거 숨기는 코드임
chrome_options.add_argument('--log-level=3')
# C드라이버에 경로 넣지 않아야함
driver = webdriver.Chrome(options=chrome_options)

path = r'크롤링/Project/02.파이썬엑셀다루기/참가자_data.xlsx'

#새로운 엑셀 파일 생성
wb = openpyxl.load_workbook(path)

ws = wb.create_sheet('주식')

codes = [
    '005930', # 삼성전자
    '000660', # SK하이닉스
    '035720'  # 카카오
]

ws.append(['종목', '현재가'])

row = 2
for code in codes:
    url = f'https://finance.naver.com/item/sise.naver?code={code}'
    driver.get(url)
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    
    # 종목
    event = soup.select_one('.wrap_company>h2>a').text #삼성전자
    
    # 현재가
    price = soup.select_one('#_nowVal').text #74,000
    price = price.replace(',', '') # 74000
    
    ws[f'A{row}'] = event
    ws[f'B{row}'] = int(price)

    row += 1 # row가 하나씩 들어남
    
    time.sleep(1)
    
wb.save(path)

print('크롤링 작업을 마쳤습니다.')
