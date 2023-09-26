from openpyxl import Workbook

wb = Workbook()

ws = wb.create_sheet('start_coding')

ws['A1'] = '스타트코딩'

ws['B1'] = 'ㅎㅇㅎㅇ'

wb.save('test.xlsx')

print('\nDvlp.H.Y.C.Sol\n')