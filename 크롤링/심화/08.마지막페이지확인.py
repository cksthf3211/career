from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup

# 크롬 드라이버 자동 업데이트
from webdriver_manager.chrome import ChromeDriverManager

from openpyxl import Workbook
from openpyxl.styles import Alignment

import time
import pyautogui
import pyperclip
import csv

# 브라우저 꺼짐 방지
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)

# 크롬창 안뜨게 함
chrome_options.add_argument('--headless') # headless 모드 활성화
chrome_options.add_argument('--disable-gpu') # GPU 가속 비활성화

# Mozilla 웹 브라우저에서 온 것처럼 인식 / 자동화된 요청을 감지하고 차단하는 것을 우회
chrome_options.add_argument("--user-agent=Mozilla/5.0")

# 불필요 메세지 없애기
chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
# 드라이버 업데이트
service = Service(executable_path=ChromeDriverManager().install())

# 옵션 적용
browser = webdriver.Chrome(service=service, options=chrome_options)

news = pyautogui.prompt('뉴스기사 입력 >>> ')
page = pyautogui.prompt('몇 페이지까지 크롤링 할까요? >>> ')

# 엑셀 생성
wb = Workbook()
ws = wb.create_sheet(news)
del wb['Sheet']

# 열 너비 조정
ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 100

print(f'{news}를 {page}페이지 검색')

# 행 번호
xlsx_num = 1

page_num = 1 #
for i in range(1, int(page) * 10, 10):
    print(f"=============== {page_num} 페이지 크롤링 중 ==============")
    
    # 웹페이지 해당 주소 이동
    path = f'https://search.naver.com/search.naver?where=news&sm=tab_jum&query={news}&start={i}'

    # url 대화
    browser.get(path)

    # 네이버에서 html 줌
    html = browser.page_source

    soup = BeautifulSoup(html, 'html.parser')

    articles = soup.select("div.info_group") # 뉴스 기사 div 10개 추출
            
    for article in articles:
        links = article.select("a.info")
        if len(links) >= 2: # 링크가 2개 이상이면
            url = links[1].attrs['href'] # 두번째 링크의 href 추출
            
            # 다시 한번 받아옴
            browser.get(url)
            news_url = browser.current_url
            html = browser.page_source
            soup_sub = BeautifulSoup(html, 'html.parser')
            
            # 연예뉴스라면 -> ? div 모양이 다름
            # 링크 앞부분에 무엇이 포함되어있는지?
            if 'entertain' in news_url:
                title = soup_sub.select_one(".end_tit")
                content = soup_sub.select_one('#articeBody')
            
            # 불필요 내용 삭제하는 기능 추가함
            elif 'sports' in news_url:
                title = soup_sub.select_one("h4.title") # 앞에 h4가 없으면 50개 나옴
                content = soup.select_one('#newsEndContents')
                # 본문 내용 안에 불필요한 내용, 공백 삭제
                divs = content.select("div")
                for div in divs:
                    div.decompose()
                paragraphs = content.select("p")
                for paragraph in paragraphs:
                    paragraph.decompose()
                    
            else:
                title = soup_sub.select_one("#title_area")
                content = soup_sub.select_one('#dic_area') # 해당 링크 본문의 아이디값 가져옴
                
            print("=============링크==========\n", news_url)
            print("=============제목==========\n", title.text.strip())
            print("=============내용==========\n", content.text.strip()) # br제거해야함
            
            ws[f'A{xlsx_num}'] = news_url
            ws[f'B{xlsx_num}'] = title.text.strip()
            ws[f'C{xlsx_num}'] = content.text.strip()
            
            # 줄바꿈
            ws[f'C{xlsx_num}'].alignment = Alignment(wrap_text=True)
            
            # 한칸 내려가기
            xlsx_num += 1
            
            time.sleep(0.7)
    # 마지막 페이지 여부 확인하기
    last_page = soup.select_one("a.btn_next").attrs['aria-disabled'] # 태그 안에 속성이 disabled인것
    if last_page == 'true':
        print('마지막 페이지입니다.')
        break
        # AttributeError: 'NoneType' object has no attribute 'attrs'
        # 이 에러는 soup가 덮어씌워져서 데이터가 없다는거임, 다르게 만들어주기.
    
    page_num += 1
    
wb.save(f'크롤링/심화/{news}_result.xlsx')
print('\nDvlp.H.Y.C.Sol\n')