import openpyxl

save_path = 'Data/03.엑셀자동화/쇼핑몰.xlsx'

# 기존 엑셀 파일 불러오기
# data_only=True 를 사용하여 수식 대신 결과값 보기
wb = openpyxl.load_workbook(save_path, data_only=True)

# data 시트 선택
ws = wb['data']

# 01. 모든 셀 데이터 가져오기
# -> 행과 열 개수를 아는 경우
# for x in range(1, 7+1):     # 행
#     for y in range(1, 5+1): # 열
#         # 출력
#         print(ws.cell(row=x, column=y).value, end=" ")
#     print() # 한 줄씩
    
# -> 행과 열 개수를 모르는 경우
# for x in range(1, ws.max_row+1):     # 행
#     for y in range(1, ws.max_column+1): # 열
#         # 출력
#         print(ws.cell(row=x, column=y).value, end=" ")
#     print() # 한 줄씩
    
# 모든 행 가져오기
# for i in ws.iter_rows():
#     print(i)
    
# 2번째 행 부터 가져오기
# -> 헤더 없애고 보기
# for i in ws.iter_rows(min_row=2):
#     print(i)

# 2번째 행 부터 5번째 행까지 가져오기
# for i in ws.iter_rows(min_row=2, max_row=5):
#     print(i)

# 2~4행, 2~4열 가져오기
for i in ws.iter_rows(min_row=2, max_row=4, min_col=2, max_col=4):
    for j in i:
        print(j.value, end=" ")
    print()